from openpyxl import load_workbook
#导入Excel文件
wb = load_workbook('a.xlsx')
#激活工作表
#ws = wb.active
#获取sheet
ws =wb['Sheet1']
#插入数据
ws['A1'] = 234
ws.cell(row=1, column=3).value = 123

#遍历数据
for row in ws.values:
    for cell in row:
        print(cell)

for row in ws.iter_rows(min_row=2, max_col=2, max_row=3):
    for cell in row:
        print(cell.value)

#保存到文件
wb.save("a.xlsx")
