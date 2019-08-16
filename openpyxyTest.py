from openpyxl import load_workbook

wb = load_workbook('aaa.xlsx')
# ws = wb['Sheet']


# 激活工作表
# ws = wb.active
ws =wb['sheet1']
ws['A1'] = 234
ws.cell(row=1, column=3).value = 123

#遍历数据
for row in ws.values:
    for cell in row:
        print(cell)

for row in ws.iter_rows(min_row=2, max_col=2, max_row=3):
    for cell in row:
        print(cell.value)


wb.save("aaa.xlsx")
